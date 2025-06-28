#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Interface web simples para visualização dos dados LGPD
Substituto do Streamlit com Flask
"""

from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import json
from datetime import datetime
import os
import asyncio
from concurrent.futures import ThreadPoolExecutor
from database import (
    obter_estatisticas,
    obter_resultados_por_dominio,
    obter_resultados_por_empresa,
    obter_dominios_unicos,
    obter_empresas_unicas,
    carregar_empresas_padrao,
    inicializar_banco,
    obter_prioridades_busca,
    inserir_prioridade_busca,
    remover_prioridade_busca,
    carregar_prioridades_padrao,
    obter_regex_patterns,
    inserir_regex_pattern,
    carregar_regex_padrao
)
import sqlite3
from main import processar_arquivos
from ai_enhanced_processor import processar_arquivos_com_ia
from database_postgresql import db_manager, initialize_postgresql
from ai_processor_simplified import initialize_simple_ai_system, process_document_simple_ai

app = Flask(__name__)
executor = ThreadPoolExecutor(max_workers=4)

# Global flag for PostgreSQL availability
POSTGRESQL_ENABLED = False

def run_async(coro):
    """Helper to run async functions in Flask context"""
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        return loop.run_until_complete(coro)
    finally:
        loop.close()

def initialize_systems():
    """Initialize both SQLite and PostgreSQL systems"""
    global POSTGRESQL_ENABLED
    
    # Initialize SQLite (existing system)
    inicializar_banco()
    
    # Try to initialize PostgreSQL system
    try:
        success = run_async(initialize_postgresql())
        if success:
            # Also initialize the AI system
            run_async(initialize_simple_ai_system())
            POSTGRESQL_ENABLED = True
            print("✅ PostgreSQL system enabled")
        else:
            print("⚠️ PostgreSQL system failed, using SQLite only")
    except Exception as e:
        print(f"⚠️ PostgreSQL initialization error: {e}")
        POSTGRESQL_ENABLED = False

@app.route('/')
def dashboard():
    """Página principal do dashboard"""
    return render_template('dashboard.html')

@app.route('/api/estatisticas')
def api_estatisticas():
    """API para obter estatísticas gerais"""
    stats = obter_estatisticas()
    return jsonify(stats)

@app.route('/api/dominios')
def api_dominios():
    """API para obter lista de domínios"""
    dominios = obter_dominios_unicos()
    return jsonify(dominios)

@app.route('/api/empresas')
def api_empresas():
    """API para obter lista de empresas"""
    empresas = obter_empresas_unicas()
    return jsonify(empresas)

@app.route('/api/resultados')
def api_resultados():
    """API para obter resultados com filtros"""
    dominio = request.args.get('dominio')
    empresa = request.args.get('empresa')
    
    if dominio and dominio != 'todos':
        resultados = obter_resultados_por_dominio(dominio)
    elif empresa and empresa != 'todos':
        resultados = obter_resultados_por_empresa(empresa)
    else:
        resultados = obter_resultados_por_dominio()
    
    return jsonify(resultados)

@app.route('/api/processar', methods=['POST'])
def api_processar():
    """API para executar processamento de arquivos"""
    try:
        data = request.get_json()
        diretorio = data.get('diretorio', 'data') if data else 'data'
        
        # Validar se o diretório existe
        if not os.path.exists(diretorio):
            return jsonify({'status': 'error', 'message': f'Diretório não encontrado: {diretorio}'})
        
        # Usar processador avançado com IA se disponível
        try:
            estatisticas = processar_arquivos_com_ia(diretorio)
            return jsonify({
                'status': 'success', 
                'message': 'Processamento com IA concluído',
                'estatisticas': estatisticas
            })
        except Exception as e:
            # Fallback para processador básico
            processar_arquivos(diretorio)
            return jsonify({'status': 'success', 'message': 'Processamento básico concluído'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/api/status-processamento')
def api_status_processamento():
    """API para obter status do processamento em tempo real"""
    # Por enquanto retorna status simples, pode ser expandido com WebSockets
    return jsonify({
        'processando': False,
        'arquivo_atual': '',
        'progresso': 0,
        'total_arquivos': 0
    })

@app.route('/api/carregar-empresas', methods=['POST'])
def api_carregar_empresas():
    """API para carregar empresas padrão"""
    try:
        carregar_empresas_padrao()
        return jsonify({'status': 'success', 'message': 'Empresas carregadas'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/api/export-excel')
def api_export_excel():
    """API para exportar dados em Excel"""
    try:
        dominio = request.args.get('dominio')
        empresa = request.args.get('empresa')
        
        if dominio and dominio != 'todos':
            resultados = obter_resultados_por_dominio(dominio)
        elif empresa and empresa != 'todos':
            resultados = obter_resultados_por_empresa(empresa)
        else:
            resultados = obter_resultados_por_dominio()
        
        if not resultados:
            return jsonify({'status': 'error', 'message': 'Nenhum dado para exportar'})
        
        # Criar DataFrame
        df = pd.DataFrame(resultados)
        
        # Definir colunas na ordem desejada
        colunas_ordenadas = [
            'dominio', 'empresa', 'tipo_dado', 'valor_encontrado',
            'arquivo_origem', 'titular_identificado', 'metodo_identificacao',
            'prioridade', 'contexto', 'data_analise', 'status_compliance'
        ]
        
        df = df[colunas_ordenadas]
        
        # Renomear colunas para português
        df.columns = [
            'Domínio', 'Empresa', 'Tipo de Dado', 'Valor Encontrado',
            'Arquivo de Origem', 'Titular Identificado', 'Método de Identificação',
            'Prioridade', 'Contexto', 'Data da Análise', 'Status de Compliance'
        ]
        
        # Salvar Excel
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'matriz_dados_lgpd_{timestamp}.xlsx'
        
        df.to_excel(filename, index=False, sheet_name='Dados LGPD')
        
        return send_file(filename, as_attachment=True, download_name=filename)
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

# === PRIORIDADE DE BUSCA APIs ===

@app.route('/api/prioridades-busca')
def api_prioridades_busca():
    """API para obter lista de prioridades de busca"""
    prioridades = obter_prioridades_busca()
    return jsonify(prioridades)

@app.route('/api/prioridades-busca', methods=['POST'])
def api_adicionar_prioridade():
    """API para adicionar nova prioridade de busca"""
    try:
        data = request.get_json()
        prioridade = data.get('prioridade')
        nome_empresa = data.get('nome_empresa')
        dominio_email = data.get('dominio_email')
        
        if not all([prioridade, nome_empresa, dominio_email]):
            return jsonify({'status': 'error', 'message': 'Todos os campos são obrigatórios'})
        
        sucesso = inserir_prioridade_busca(prioridade, nome_empresa, dominio_email)
        
        if sucesso:
            return jsonify({'status': 'success', 'message': 'Prioridade adicionada com sucesso'})
        else:
            return jsonify({'status': 'error', 'message': 'Erro ao adicionar prioridade'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/api/prioridades-busca/<int:prioridade_id>', methods=['DELETE'])
def api_remover_prioridade(prioridade_id):
    """API para remover prioridade de busca"""
    try:
        sucesso = remover_prioridade_busca(prioridade_id)
        
        if sucesso:
            return jsonify({'status': 'success', 'message': 'Prioridade removida com sucesso'})
        else:
            return jsonify({'status': 'error', 'message': 'Erro ao remover prioridade'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/api/carregar-prioridades-padrao', methods=['POST'])
def api_carregar_prioridades_padrao():
    """API para carregar prioridades padrão"""
    try:
        carregar_prioridades_padrao()
        return jsonify({'status': 'success', 'message': 'Prioridades padrão carregadas'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

# === REGEX PATTERNS APIs ===

@app.route('/api/regex-patterns')
def api_regex_patterns():
    """API para obter lista de padrões regex"""
    patterns = obter_regex_patterns()
    return jsonify(patterns)

@app.route('/api/regex-patterns', methods=['POST'])
def api_adicionar_regex():
    """API para adicionar novo padrão regex"""
    try:
        data = request.get_json()
        nome_campo = data.get('nome_campo')
        pattern_regex = data.get('pattern_regex')
        explicacao = data.get('explicacao', '')
        
        if not all([nome_campo, pattern_regex]):
            return jsonify({'status': 'error', 'message': 'Nome do campo e padrão são obrigatórios'})
        
        sucesso = inserir_regex_pattern(nome_campo, pattern_regex, explicacao)
        
        if sucesso:
            return jsonify({'status': 'success', 'message': 'Padrão regex adicionado com sucesso'})
        else:
            return jsonify({'status': 'error', 'message': 'Erro ao adicionar padrão regex'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/api/carregar-regex-padrao', methods=['POST'])
def api_carregar_regex_padrao():
    """API para carregar padrões regex padrão"""
    try:
        carregar_regex_padrao()
        return jsonify({'status': 'success', 'message': 'Padrões regex padrão carregados'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

# === NAVEGAÇÃO DE DIRETÓRIOS ===

@app.route('/api/listar-diretorios')
def api_listar_diretorios():
    """API para listar estrutura de diretórios"""
    try:
        caminho = request.args.get('caminho', '.')
        
        # Garantir que o caminho é seguro e dentro dos limites
        if not caminho.startswith('.') and not caminho.startswith('/opt/privacy'):
            caminho = os.path.join('.', caminho)
        
        # Normalizar o caminho
        caminho_absoluto = os.path.abspath(caminho)
        
        # Verificar se o diretório existe
        if not os.path.exists(caminho_absoluto):
            return jsonify({
                'status': 'error', 
                'message': f'Diretório não encontrado: {caminho}'
            })
        
        # Listar conteúdo do diretório
        items = []
        try:
            for item in sorted(os.listdir(caminho_absoluto)):
                if item.startswith('.'):  # Ignorar arquivos/pastas ocultos
                    continue
                    
                item_path = os.path.join(caminho_absoluto, item)
                is_dir = os.path.isdir(item_path)
                
                # Para diretórios, contar arquivos suportados
                file_count = 0
                if is_dir:
                    try:
                        from file_scanner import listar_arquivos_recursivos
                        arquivos = listar_arquivos_recursivos(item_path)
                        file_count = len(arquivos) if arquivos else 0
                    except:
                        file_count = 0
                
                items.append({
                    'nome': item,
                    'caminho': os.path.relpath(item_path, '.'),
                    'tipo': 'diretorio' if is_dir else 'arquivo',
                    'tamanho': 0 if is_dir else os.path.getsize(item_path),
                    'arquivos_suportados': file_count if is_dir else 0,
                    'modificado': os.path.getmtime(item_path)
                })
        
        except PermissionError:
            return jsonify({
                'status': 'error',
                'message': f'Sem permissão para acessar: {caminho}'
            })
        
        return jsonify({
            'status': 'success',
            'caminho_atual': os.path.relpath(caminho_absoluto, '.'),
            'items': items
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Erro ao listar diretórios: {str(e)}'
        })

@app.route('/api/validar-diretorio')
def api_validar_diretorio():
    """API para validar se um diretório contém arquivos processáveis"""
    try:
        caminho = request.args.get('caminho', 'data')
        
        # Normalizar o caminho
        if not caminho.startswith('.') and not caminho.startswith('/opt/privacy'):
            caminho = os.path.join('.', caminho)
        
        caminho_absoluto = os.path.abspath(caminho)
        
        if not os.path.exists(caminho_absoluto):
            return jsonify({
                'status': 'error',
                'valido': False,
                'message': f'Diretório não encontrado: {caminho}'
            })
        
        if not os.path.isdir(caminho_absoluto):
            return jsonify({
                'status': 'error',
                'valido': False,
                'message': f'Caminho não é um diretório: {caminho}'
            })
        
        # Verificar arquivos suportados
        try:
            from file_scanner import listar_arquivos_recursivos
            arquivos = listar_arquivos_recursivos(caminho_absoluto)
            total_arquivos = len(arquivos) if arquivos else 0
            
            return jsonify({
                'status': 'success',
                'valido': True,
                'total_arquivos': total_arquivos,
                'caminho': os.path.relpath(caminho_absoluto, '.'),
                'message': f'Diretório válido com {total_arquivos} arquivos processáveis'
            })
            
        except Exception as scan_error:
            return jsonify({
                'status': 'warning',
                'valido': True,
                'total_arquivos': 0,
                'message': f'Diretório acessível mas erro ao escanear: {str(scan_error)}'
            })
            
    except Exception as e:
        return jsonify({
            'status': 'error',
            'valido': False,
            'message': f'Erro ao validar diretório: {str(e)}'
        })

# === GERENCIAMENTO DE INCIDENTES ===

@app.route('/api/incidentes', methods=['GET'])
def api_listar_incidentes():
    """API para listar incidentes"""
    try:
        if POSTGRESQL_ENABLED:
            conn = run_async(db_manager.get_connection())
            results = run_async(conn.fetch('''
                SELECT id, empresa, data_incidente, tipo_incidente, descricao, 
                       status, criado_em, atualizado_em
                FROM incidentes_lgpd 
                ORDER BY criado_em DESC
            '''))
            run_async(conn.close())
            
            incidentes = []
            for row in results:
                incidentes.append({
                    'id': row['id'],
                    'empresa': row['empresa'],
                    'data_incidente': row['data_incidente'].strftime('%Y-%m-%d') if row['data_incidente'] else '',
                    'tipo_incidente': row['tipo_incidente'],
                    'descricao': row['descricao'],
                    'status': row['status'],
                    'criado_em': row['criado_em'].strftime('%Y-%m-%d %H:%M') if row['criado_em'] else '',
                    'atualizado_em': row['atualizado_em'].strftime('%Y-%m-%d %H:%M') if row['atualizado_em'] else ''
                })
            
            return jsonify(incidentes)
        else:
            # Fallback para SQLite
            from database import obter_conexao
            conn = obter_conexao()
            cursor = conn.cursor()
            
            # Criar tabela se não existir
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS incidentes_lgpd (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    empresa TEXT NOT NULL,
                    data_incidente DATE,
                    tipo_incidente TEXT NOT NULL,
                    descricao TEXT,
                    status TEXT DEFAULT 'Aberto',
                    criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            cursor.execute('''
                SELECT id, empresa, data_incidente, tipo_incidente, descricao, 
                       status, criado_em, atualizado_em
                FROM incidentes_lgpd 
                ORDER BY criado_em DESC
            ''')
            
            results = cursor.fetchall()
            conn.close()
            
            incidentes = []
            for row in results:
                incidentes.append({
                    'id': row[0],
                    'empresa': row[1],
                    'data_incidente': row[2] or '',
                    'tipo_incidente': row[3],
                    'descricao': row[4],
                    'status': row[5],
                    'criado_em': row[6] or '',
                    'atualizado_em': row[7] or ''
                })
            
            return jsonify(incidentes)
            
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Erro ao listar incidentes: {str(e)}'})

@app.route('/api/incidentes', methods=['POST'])
def api_criar_incidente():
    """API para criar novo incidente"""
    try:
        data = request.get_json()
        empresa = data.get('empresa', '').strip()
        data_incidente = data.get('data_incidente', '')
        tipo_incidente = data.get('tipo_incidente', '').strip()
        descricao = data.get('descricao', '').strip()
        
        if not all([empresa, tipo_incidente, descricao]):
            return jsonify({
                'status': 'error', 
                'message': 'Empresa, tipo e descrição são obrigatórios'
            })
        
        if POSTGRESQL_ENABLED:
            conn = run_async(db_manager.get_connection())
            
            # Criar tabela se não existir
            await_result = run_async(conn.execute('''
                CREATE TABLE IF NOT EXISTS incidentes_lgpd (
                    id SERIAL PRIMARY KEY,
                    empresa VARCHAR(255) NOT NULL,
                    data_incidente DATE,
                    tipo_incidente VARCHAR(100) NOT NULL,
                    descricao TEXT,
                    status VARCHAR(50) DEFAULT 'Aberto',
                    criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            '''))
            
            # Inserir incidente
            incident_id = run_async(conn.fetchval('''
                INSERT INTO incidentes_lgpd (empresa, data_incidente, tipo_incidente, descricao)
                VALUES ($1, $2, $3, $4)
                RETURNING id
            ''', empresa, data_incidente or None, tipo_incidente, descricao))
            
            run_async(conn.close())
            
            return jsonify({
                'status': 'success',
                'message': 'Incidente registrado com sucesso',
                'id': incident_id
            })
        else:
            # Fallback para SQLite
            from database import obter_conexao
            conn = obter_conexao()
            cursor = conn.cursor()
            
            # Criar tabela se não existir
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS incidentes_lgpd (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    empresa TEXT NOT NULL,
                    data_incidente DATE,
                    tipo_incidente TEXT NOT NULL,
                    descricao TEXT,
                    status TEXT DEFAULT 'Aberto',
                    criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Inserir incidente
            cursor.execute('''
                INSERT INTO incidentes_lgpd (empresa, data_incidente, tipo_incidente, descricao)
                VALUES (?, ?, ?, ?)
            ''', (empresa, data_incidente or None, tipo_incidente, descricao))
            
            incident_id = cursor.lastrowid
            conn.commit()
            conn.close()
            
            return jsonify({
                'status': 'success',
                'message': 'Incidente registrado com sucesso',
                'id': incident_id
            })
            
    except Exception as e:
        return jsonify({
            'status': 'error', 
            'message': f'Erro ao criar incidente: {str(e)}'
        })

@app.route('/api/incidentes/<int:incident_id>', methods=['PUT'])
def api_atualizar_incidente(incident_id):
    """API para atualizar status de incidente"""
    try:
        data = request.get_json()
        status = data.get('status', '').strip()
        
        if not status:
            return jsonify({'status': 'error', 'message': 'Status é obrigatório'})
        
        if POSTGRESQL_ENABLED:
            conn = run_async(db_manager.get_connection())
            
            run_async(conn.execute('''
                UPDATE incidentes_lgpd 
                SET status = $1, atualizado_em = CURRENT_TIMESTAMP
                WHERE id = $2
            ''', status, incident_id))
            
            run_async(conn.close())
        else:
            # Fallback para SQLite
            from database import obter_conexao
            conn = obter_conexao()
            cursor = conn.cursor()
            
            cursor.execute('''
                UPDATE incidentes_lgpd 
                SET status = ?, atualizado_em = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (status, incident_id))
            
            conn.commit()
            conn.close()
        
        return jsonify({
            'status': 'success',
            'message': 'Status do incidente atualizado'
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Erro ao atualizar incidente: {str(e)}'
        })

# New AI-powered endpoints for PostgreSQL system
@app.route('/api/ai-metrics')
def api_ai_metrics():
    """API para métricas do sistema de IA em tempo real"""
    if not POSTGRESQL_ENABLED:
        return jsonify({'status': 'disabled', 'message': 'Sistema PostgreSQL não disponível'})
    
    try:
        metrics = run_async(db_manager.get_dashboard_metrics())
        return jsonify(metrics)
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/api/processing-queue')
def api_processing_queue():
    """API para fila de processamento com prioridades AI"""
    if not POSTGRESQL_ENABLED:
        return jsonify([])
    
    try:
        queue = run_async(db_manager.get_processing_queue())
        return jsonify(queue)
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/api/search-priorities-pg')
def api_search_priorities_pg():
    """API para prioridades de busca do PostgreSQL"""
    if not POSTGRESQL_ENABLED:
        return jsonify([])
    
    try:
        priorities = run_async(db_manager.get_search_priorities())
        return jsonify(priorities)
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/api/system-status')
def api_system_status():
    """API para status dos sistemas"""
    return jsonify({
        'sqlite_enabled': True,
        'postgresql_enabled': POSTGRESQL_ENABLED,
        'ai_system_enabled': POSTGRESQL_ENABLED,
        'timestamp': datetime.now().isoformat()
    })

# === RELATÓRIOS E EXPORTAÇÃO EXCEL ===

@app.route('/api/export-excel-filtrado', methods=['POST'])
def api_export_excel_filtrado():
    """API para exportar dados filtrados para Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import io
        from flask import send_file
        
        # Obter filtros da requisição
        filtros = request.get_json() if request.is_json else {}
        
        # Aplicar filtros aos dados
        dados_filtrados = aplicar_filtros_relatorio(filtros)
        
        # Criar workbook do Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Dados LGPD"
        
        # Definir cabeçalhos
        cabecalhos = [
            'ID', 'Arquivo', 'Titular', 'Tipo de Dado', 'Valor', 
            'Contexto', 'Prioridade', 'Método Identificação', 'Timestamp',
            'Empresa Identificada', 'Domínio Email', 'Formato Arquivo'
        ]
        
        # Estilizar cabeçalhos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, cabecalho in enumerate(cabecalhos, 1):
            cell = ws.cell(row=1, column=col, value=cabecalho)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Adicionar dados
        for row_idx, dado in enumerate(dados_filtrados, 2):
            ws.cell(row=row_idx, column=1, value=dado.get('id', ''))
            ws.cell(row=row_idx, column=2, value=dado.get('arquivo', ''))
            ws.cell(row=row_idx, column=3, value=dado.get('titular', ''))
            ws.cell(row=row_idx, column=4, value=dado.get('campo', ''))
            ws.cell(row=row_idx, column=5, value=dado.get('valor', ''))
            ws.cell(row=row_idx, column=6, value=dado.get('contexto', ''))
            ws.cell(row=row_idx, column=7, value=dado.get('prioridade', ''))
            ws.cell(row=row_idx, column=8, value=dado.get('origem_identificacao', ''))
            ws.cell(row=row_idx, column=9, value=dado.get('timestamp', ''))
            ws.cell(row=row_idx, column=10, value=extrair_empresa_do_contexto(dado.get('contexto', '')))
            ws.cell(row=row_idx, column=11, value=extrair_dominio_email(dado.get('valor', '')))
            ws.cell(row=row_idx, column=12, value=obter_extensao_arquivo(dado.get('arquivo', '')))
        
        # Ajustar largura das colunas
        for col in range(1, len(cabecalhos) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15
        
        # Criar segunda aba com estatísticas
        ws_stats = wb.create_sheet(title="Estatísticas")
        estatisticas = gerar_estatisticas_relatorio(dados_filtrados)
        
        # Adicionar estatísticas
        ws_stats.cell(row=1, column=1, value="Estatísticas do Relatório LGPD").font = Font(bold=True, size=14)
        row = 3
        
        for categoria, valores in estatisticas.items():
            ws_stats.cell(row=row, column=1, value=categoria).font = Font(bold=True)
            row += 1
            
            for chave, valor in valores.items():
                ws_stats.cell(row=row, column=2, value=chave)
                ws_stats.cell(row=row, column=3, value=valor)
                row += 1
            row += 1
        
        # Criar terceira aba com resumo por empresa
        ws_empresa = wb.create_sheet(title="Resumo por Empresa")
        resumo_empresas = gerar_resumo_empresas(dados_filtrados)
        
        cabecalhos_empresa = ['Empresa', 'Total de Registros', 'CPFs', 'Emails', 'Telefones', 'Prioridade Média']
        for col, cabecalho in enumerate(cabecalhos_empresa, 1):
            cell = ws_empresa.cell(row=1, column=col, value=cabecalho)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for row_idx, (empresa, dados) in enumerate(resumo_empresas.items(), 2):
            ws_empresa.cell(row=row_idx, column=1, value=empresa)
            ws_empresa.cell(row=row_idx, column=2, value=dados['total'])
            ws_empresa.cell(row=row_idx, column=3, value=dados['cpfs'])
            ws_empresa.cell(row=row_idx, column=4, value=dados['emails'])
            ws_empresa.cell(row=row_idx, column=5, value=dados['telefones'])
            ws_empresa.cell(row=row_idx, column=6, value=dados['prioridade_media'])
        
        # Salvar arquivo em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"relatorio_lgpd_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Erro na exportação: {str(e)}'})

@app.route('/api/relatorio-completo', methods=['POST'])
def api_relatorio_completo():
    """API para gerar relatório completo com análises avançadas"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, PieChart, Reference
        import io
        
        filtros = request.get_json() if request.is_json else {}
        dados_filtrados = aplicar_filtros_relatorio(filtros)
        
        wb = Workbook()
        
        # Aba 1: Dashboard Executivo
        ws_dashboard = wb.active
        ws_dashboard.title = "Dashboard Executivo"
        
        # Criar dashboard executivo simples
        ws_dashboard.cell(row=1, column=1, value="Dashboard Executivo LGPD").font = Font(bold=True, size=16)
        ws_dashboard.cell(row=3, column=1, value=f"Total de Registros: {len(dados_filtrados)}")
        ws_dashboard.cell(row=4, column=1, value=f"Data do Relatório: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        
        # Aba 2: Dados Detalhados  
        ws_detalhes = wb.create_sheet(title="Dados Detalhados")
        cabecalhos = ['ID', 'Arquivo', 'Titular', 'Tipo', 'Valor', 'Contexto', 'Prioridade']
        for col, cabecalho in enumerate(cabecalhos, 1):
            ws_detalhes.cell(row=1, column=col, value=cabecalho).font = Font(bold=True)
        
        for row_idx, dado in enumerate(dados_filtrados, 2):
            ws_detalhes.cell(row=row_idx, column=1, value=dado.get('id', ''))
            ws_detalhes.cell(row=row_idx, column=2, value=dado.get('arquivo', ''))
            ws_detalhes.cell(row=row_idx, column=3, value=dado.get('titular', ''))
            ws_detalhes.cell(row=row_idx, column=4, value=dado.get('campo', ''))
            ws_detalhes.cell(row=row_idx, column=5, value=dado.get('valor', ''))
            ws_detalhes.cell(row=row_idx, column=6, value=dado.get('contexto', ''))
            ws_detalhes.cell(row=row_idx, column=7, value=dado.get('prioridade', ''))
        
        # Salvar arquivo
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"relatorio_completo_lgpd_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Erro no relatório completo: {str(e)}'})

def aplicar_filtros_relatorio(filtros):
    """Aplica filtros aos dados para relatórios"""
    conn = sqlite3.connect('lgpd_data.db')
    cursor = conn.cursor()
    
    # Query base
    query = """
    SELECT id, arquivo, titular, campo, valor, contexto, prioridade, 
           origem_identificacao, timestamp
    FROM dados_extraidos WHERE 1=1
    """
    
    parametros = []
    
    # Aplicar filtros
    if filtros.get('dominio') and filtros['dominio'] != 'todos':
        query += " AND valor LIKE ?"
        parametros.append(f"%{filtros['dominio']}%")
    
    if filtros.get('empresa') and filtros['empresa'] != 'todos':
        query += " AND (titular LIKE ? OR contexto LIKE ?)"
        parametros.extend([f"%{filtros['empresa']}%", f"%{filtros['empresa']}%"])
    
    if filtros.get('prioridade') and filtros['prioridade'] != 'todos':
        query += " AND prioridade = ?"
        parametros.append(filtros['prioridade'])
    
    if filtros.get('tipo_dado') and filtros['tipo_dado'] != 'todos':
        query += " AND campo = ?"
        parametros.append(filtros['tipo_dado'])
    
    if filtros.get('data_inicial'):
        query += " AND DATE(timestamp) >= ?"
        parametros.append(filtros['data_inicial'])
    
    if filtros.get('data_final'):
        query += " AND DATE(timestamp) <= ?"
        parametros.append(filtros['data_final'])
    
    if filtros.get('origem_identificacao') and filtros['origem_identificacao'] != 'todos':
        query += " AND origem_identificacao = ?"
        parametros.append(filtros['origem_identificacao'])
    
    if filtros.get('formato_arquivo') and filtros['formato_arquivo'] != 'todos':
        query += " AND arquivo LIKE ?"
        parametros.append(f"%{filtros['formato_arquivo']}")
    
    if filtros.get('com_contexto'):
        query += " AND contexto IS NOT NULL AND contexto != ''"
    
    query += " ORDER BY timestamp DESC"
    
    cursor.execute(query, parametros)
    resultados = cursor.fetchall()
    
    # Converter para lista de dicionários
    colunas = ['id', 'arquivo', 'titular', 'campo', 'valor', 'contexto', 
               'prioridade', 'origem_identificacao', 'timestamp']
    
    dados = []
    for resultado in resultados:
        dados.append(dict(zip(colunas, resultado)))
    
    conn.close()
    return dados

def extrair_empresa_do_contexto(contexto):
    """Extrai nome da empresa do contexto"""
    if not contexto:
        return ""
    
    # Lista de empresas conhecidas para identificação
    empresas_conhecidas = [
        'BRADESCO', 'PETROBRAS', 'ONS', 'EMBRAER', 'REDE DOR', 'ED GLOBO', 
        'GLOBO', 'ELETROBRAS', 'CREFISA', 'EQUINIX', 'COHESITY', 'NETAPP', 
        'HITACHI', 'LENOVO', 'MICROSOFT', 'GOOGLE', 'AMAZON', 'IBM'
    ]
    
    contexto_upper = contexto.upper()
    for empresa in empresas_conhecidas:
        if empresa in contexto_upper:
            return empresa
    
    return "Não Identificada"

def extrair_dominio_email(valor):
    """Extrai domínio de email do valor"""
    if '@' in valor:
        return valor.split('@')[1] if len(valor.split('@')) > 1 else ""
    return ""

def obter_extensao_arquivo(arquivo):
    """Obtém extensão do arquivo"""
    import os
    return os.path.splitext(arquivo)[1] if arquivo else ""

def gerar_estatisticas_relatorio(dados):
    """Gera estatísticas do relatório"""
    total_registros = len(dados)
    
    # Contar por tipo de dado
    tipos_dados = {}
    prioridades = {}
    origens = {}
    
    for dado in dados:
        # Tipos de dados
        tipo = dado.get('campo', 'Não identificado')
        tipos_dados[tipo] = tipos_dados.get(tipo, 0) + 1
        
        # Prioridades
        prioridade = dado.get('prioridade', 'Não definida')
        prioridades[prioridade] = prioridades.get(prioridade, 0) + 1
        
        # Origens de identificação
        origem = dado.get('origem_identificacao', 'Não definida')
        origens[origem] = origens.get(origem, 0) + 1
    
    return {
        'Resumo Geral': {
            'Total de Registros': total_registros,
            'Registros com Alta Prioridade': prioridades.get('Alta', 0),
            'Registros com Contexto': len([d for d in dados if d.get('contexto')])
        },
        'Por Tipo de Dado': tipos_dados,
        'Por Prioridade': prioridades,
        'Por Origem de Identificação': origens
    }

def gerar_resumo_empresas(dados):
    """Gera resumo por empresa"""
    empresas = {}
    
    for dado in dados:
        empresa = extrair_empresa_do_contexto(dado.get('contexto', ''))
        if empresa not in empresas:
            empresas[empresa] = {
                'total': 0,
                'cpfs': 0,
                'emails': 0,
                'telefones': 0,
                'prioridades': []
            }
        
        empresas[empresa]['total'] += 1
        
        tipo_dado = dado.get('campo', '').lower()
        if 'cpf' in tipo_dado:
            empresas[empresa]['cpfs'] += 1
        elif 'email' in tipo_dado:
            empresas[empresa]['emails'] += 1
        elif 'telefone' in tipo_dado:
            empresas[empresa]['telefones'] += 1
        
        prioridade = dado.get('prioridade', '')
        if prioridade:
            empresas[empresa]['prioridades'].append(prioridade)
    
    # Calcular prioridade média
    for empresa in empresas:
        prioridades = empresas[empresa]['prioridades']
        if prioridades:
            valores_prioridade = {'Alta': 3, 'Média': 2, 'Baixa': 1}
            media = sum(valores_prioridade.get(p, 1) for p in prioridades) / len(prioridades)
            empresas[empresa]['prioridade_media'] = round(media, 2)
        else:
            empresas[empresa]['prioridade_media'] = 0
    
    return empresas

@app.route('/api/test-ai', methods=['POST'])
def test_ai_integration():
    """Endpoint para testar integração OpenAI"""
    try:
        from openai import OpenAI
        import os
        
        data = request.get_json()
        texto = data.get('texto', 'Teste de funcionalidade OpenAI')
        
        client = OpenAI(api_key=os.environ.get('OPENAI_API_KEY'))
        
        response = client.chat.completions.create(
            model="gpt-3.5-turbo-1106",
            messages=[
                {
                    "role": "system", 
                    "content": "Você é um especialista em LGPD. Analise o texto e identifique dados pessoais, classifique a sensibilidade e responda em JSON."
                },
                {
                    "role": "user", 
                    "content": f"Analise este texto: {texto}"
                }
            ],
            response_format={"type": "json_object"},
            max_tokens=500
        )
        
        result = {
            'status': 'success',
            'openai_active': True,
            'model_used': 'gpt-3.5-turbo-1106',
            'analysis': response.choices[0].message.content,
            'tokens_used': response.usage.total_tokens if response.usage else 0
        }
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'openai_active': False,
            'error': str(e)
        }), 500

@app.route('/health')
def health_check():
    """Health check endpoint para load balancer"""
    return "healthy\n", 200, {'Content-Type': 'text/plain'}

if __name__ == '__main__':
    # Inicializar sistemas
    initialize_systems()
    
    # Verificar se existe pasta templates
    if not os.path.exists('templates'):
        os.makedirs('templates')
    
    print("🚀 Iniciando servidor web na porta 5000...")
    print(f"📊 Sistema SQLite: Ativo")
    print(f"🤖 Sistema PostgreSQL: {'Ativo' if POSTGRESQL_ENABLED else 'Inativo'}")
    
    app.run(host='0.0.0.0', port=5000, debug=True)